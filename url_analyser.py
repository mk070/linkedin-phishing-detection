import requests
import pandas as pd
import re
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup

def check_url_availability(url):
    """Checks if the URL is accessible by sending a request."""
    try:
        response = requests.get(url, allow_redirects=True, timeout=5)
        return response.status_code in [200, 301, 302]  # Valid statuses
    except requests.exceptions.RequestException:
        return False  # URL is not accessible

def classify_vulnerability(score):
    """Classifies the vulnerability based on the total score."""
    if score <= 0:
        return "No risk"
    elif score == 1:
        return "Low"
    elif score <=3:
        return "Medium"
    else: 
        return "Critical"


def rule3(url):
    # Define the dataset of phishing keywords inside the function
    keywords = [
        # Account Management
        "login", "secure", "account", "signin", "credentials", "password", "reset", "update", "confirm", "verification", "safety", "suspicious",

        # Urgency and Offers
        "free", "gift", "prize", "urgent", "immediate", "act now", "limited time", "exclusive",

        # Action Words
        "click", "here", "download", "access", "install", "open", "submit", "request", "claim", "verify",

        # Financial and Transactional Terms
        "banking", "payment", "invoice", "shipping", "transaction", "refund", "transfer", "credit", "debit",

        # Miscellaneous
        "support", "customer service", "update your info", "account activity", "security alert", "breach", "suspicious activity",
        "unauthorized", "reset your password", "new message", "terms and conditions", "click to win", "winner",
    ]

    # Check the URL against the keywords
    matched = 0
    for keyword in keywords:
        if keyword in url:
            matched = 1
            break  # Stop checking further if a match is found

    return matched  # Return score (1 for match, 0 for no match)

def rule1(url):
 
  def check_url_availability(url):
    """Checks if the URL is accessible by sending a request."""
    try:
        response = requests.get(url, allow_redirects=True, timeout=5)
        return response.status_code in [200, 301, 302]  # Valid statuses
    except requests.exceptions.RequestException:
        return False  # URL is not accessible

  def check_url_indexing(url):
      """Checks if a URL is indexed by search engines."""
      parsed_url = urlparse(url)
      domain = parsed_url.netloc
      search_engines = [
          f"https://www.google.com/search?q=site:{domain}",
          f"https://search.yahoo.com/search?p=site:{domain}",
          f"https://www.bing.com/search?q=site:{domain}"
      ]
      
      found_in_all = True  # Assume found until proven otherwise

      for engine in search_engines:
          try:
              response = requests.get(engine, headers={'User-Agent': 'Mozilla/5.0'}, timeout=5)
              if response.status_code == 200:
                  # Check if domain is found in the response text
                  if re.search(r'\b' + re.escape(domain) + r'\b', response.text):
                      # URL found in this search engine
                      continue
                  else:
                      found_in_all = False  # URL not found in this search engine
              else:
                  found_in_all = False  # Error in search engine check
          except requests.exceptions.RequestException:
              found_in_all = False  # Error in search engine check
      
      return 0 if found_in_all else 1  # 0 for safe (found in all), 1 for phishing


  is_accessible = check_url_availability(url)
  if not is_accessible:
      return 1  # Not accessible, consider as phishing

  return check_url_indexing(url)

def rule2(url):
  import requests
  from bs4 import BeautifulSoup
  from urllib.parse import urlparse

  def analyze_url(url):
      try:
          response = requests.get(url)
          response.raise_for_status()  # Raise an exception for bad status codes
          
          soup = BeautifulSoup(response.content, 'html.parser')

          # Extracting domain name
          parsed_url = urlparse(url)
          domain = parsed_url.netloc

          # Extract the title (optional)
          title = soup.find('title').text if soup.find('title') else "No title tag"
          return 0
          

      except requests.exceptions.RequestException:
          # If there's an error (e.g., URL not found), print "URL not found" and mark result as 1
          return 1
  return analyze_url(url)

#Rule 5: IF a URL contains any of the following characters [-, _, 0-9, @, “,”, ;] OR contains a non-standard port, THEN the webpage is potentially phishing.
def rule5(url):
    # Check for special characters in the domain or path
    special_characters_pattern = r'[_0-9@“,”;!+%]'  # Added additional characters
    non_standard_port_pattern = r':(?!80|443)(\d+)'  # Matches non-standard ports
    ip_address_pattern = r'http://\d{1,3}(?:\.\d{1,3}){3}(:\d+)?'  # Matches IP address URLs
    
    # Check for IP addresses
    if re.match(ip_address_pattern, url):
        return 1  # Potential phishing URL (IP address)
    
    # Check for special characters or non-standard ports
    if re.search(special_characters_pattern, url) or re.search(non_standard_port_pattern, url):
        return 1  # Potential phishing URL
    
    return 0  # Likely safe URL

#Rule 4: IF a webpage‟s URL is IP based (hex-based, octal, or decimal-based), THEN the webpage is potentially a phishing attack.
def rule4(url):

    """
    Check if the given URL is potentially a phishing attack based on IP address detection.
    
    Args:
        url (str): The URL to check.
    
    Returns:
        int: Score indicating if the URL is potentially phishing (1) or safe (0).
    """
    
    # Remove protocol and trailing slashes from the URL
    url = url.split("//")[-1].split("/")[0].strip()
    
    # Check for decimal IP addresses (e.g., 192.168.1.1)
    decimal_ip_pattern = re.compile(
        r'^((25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
    )

    # Check for hexadecimal IP addresses (e.g., 0xC0A80101)
    hex_ip_pattern = re.compile(r'^0x[0-9A-Fa-f]{8}$')
    
    # Check for octal IP addresses (e.g., 0300.0250.0001.0001)
    octal_ip_pattern = re.compile(r'^(0[0-7]{3}\.){3}0[0-7]{3}$')
    
    # Check if the URL matches any of the patterns
    is_decimal_ip = bool(decimal_ip_pattern.match(url))
    is_hex_ip = bool(hex_ip_pattern.match(url))
    is_octal_ip = bool(octal_ip_pattern.match(url))
    
    # Return 1 for phishing (IP-based) and 0 for safe URLs
    return 2 if is_decimal_ip or is_hex_ip or is_octal_ip else 0


def rule7(url):
    import requests
    import json

    # Replace this with your Google Safe Browsing API key
    API_KEY = 'AIzaSyDecqPBmR2_-U-PZA6KHf7Ok6tLmmMASmI'

    def check_url_with_google_safe_browsing(url):
        api_url = f'https://safebrowsing.googleapis.com/v4/threatMatches:find?key={API_KEY}'

        # Payload to send to the Google Safe Browsing API
        payload = {
            "client": {
                "clientId": "yourcompanyname",
                "clientVersion": "1.5.2"
            },
            "threatInfo": {
                "threatTypes": ["MALWARE", "SOCIAL_ENGINEERING"],
                "platformTypes": ["ANY_PLATFORM"],
                "threatEntryTypes": ["URL"],
                "threatEntries": [
                    {"url": url}
                ]
            }
        }

        try:
            response = requests.post(api_url, json=payload)
            response.raise_for_status()
            result = response.json()

            # Check if any threats were found
            if "matches" in result:
                return 4  # URL is blacklisted
            else:
                return 0  # URL is safe

        except requests.RequestException as e:
            print(f"Error: {e}")
            return -1  # Return -1 for an API or network error

    
    result = check_url_with_google_safe_browsing(url)
    return result

#Rule 11: IF a webpage contains META tag AND the refresh property‟s destination URL is in external domain OR it belongs to a blacklist, THEN the webpage is potentially phishing.
def rule11(url):
  def check_meta_refresh(url):
        """Checks for a meta refresh tag in the HTML."""
        try:
            response = requests.get(url, timeout=5)
            soup = BeautifulSoup(response.content, 'html.parser')
            meta_refresh = soup.find('meta', attrs={'http-equiv': 'refresh'})
            if meta_refresh:
                return 1  # Meta refresh tag found, potentially phishing
            return 0  # No meta refresh tag found
        except requests.exceptions.RequestException:
            return 2  # Error in accessing the URL, flag as phishing
  return check_meta_refresh(url)

def rule14(url):
    import requests
    from bs4 import BeautifulSoup
    from urllib.parse import urlparse, urljoin

    def is_internal_link(url, base_url):
        """Checks if a link is internal by comparing its base domain."""
        parsed_base_url = urlparse(base_url)
        parsed_url = urlparse(url)
        return parsed_url.netloc == parsed_base_url.netloc or parsed_url.netloc == ''

    def analyze_webpage(url):
        # Ensure the URL has a scheme
        if not urlparse(url).scheme:
            url = "http://" + url  # Default to HTTP if no scheme is provided

        try:
            # Fetch the HTML content
            response = requests.get(url)
            response.raise_for_status()  # Raise an error for bad responses

            # Parse HTML content
            soup = BeautifulSoup(response.text, 'html.parser')

            # Step 1: Check for password input fields
            password_fields = soup.find_all('input', {'type': 'password'})
            contains_password_field = len(password_fields) > 0

            # Step 2: Count internal and external links
            internal_links = 0
            external_links = 0
            base_url = urlparse(url).scheme + "://" + urlparse(url).netloc

            for link in soup.find_all('a', href=True):
                href = link['href']
                full_url = urljoin(base_url, href)
                if is_internal_link(full_url, base_url):
                    internal_links += 1
                else:
                    external_links += 1

            # Step 3: Declare the webpage as potentially phishing or not
            if contains_password_field and external_links > internal_links:
                return 3
            else:
                return 0
        
        except Exception:  # Catch any exception
            return 0  # Classify as phishing for any error

    # Test the function
    # Replace with the actual URL to test
    result = analyze_webpage(url)
    return result

def analyze_urls(urls):
    """Analyzes a list of URLs and returns a DataFrame with URLs and their Rule scores."""
    results = []
    total_score = 0
    for url in urls:
        is_accessible = check_url_availability(url)
        Rule1_Score = rule1(url) if is_accessible else 1 
        Rule2_Score = rule2(url) 
        Rule3_Score = rule3(url) 
        Rule4_Score = rule4(url) 
        Rule5_Score = rule5(url)
        Rule7_Score = rule7(url)
        Rule11_Score = rule11(url)
        Rule14_Score = rule14(url)
        total_score = Rule1_Score + Rule2_Score + Rule3_Score + Rule4_Score + Rule5_Score + Rule7_Score + Rule14_Score+Rule11_Score
        vulnerability = classify_vulnerability(total_score)
        results.append({
            "URL": url,
            "Redirect-Registration": Rule1_Score,
            "Domain_validation": Rule2_Score,
            "Suspicious_keyword": Rule3_Score,
            "ip-address_validation": Rule4_Score,
            "unwanted_special_characters": Rule5_Score,
            "google_api_validation": Rule7_Score,
            "external_password_validation": Rule14_Score,
            "meta_referesh":Rule11_Score,
            "total_score": total_score,
            "Vulnerability": vulnerability
        })

    return pd.DataFrame(results)


def apply_color_to_vulnerability_column(excel_file, sheet_name):
    """Applies color coding to the vulnerability column based on the vulnerability level."""
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    # Define the colors for the vulnerability levels
    color_mapping = {
        "Low": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),  # Green
        "Medium": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Yellow
        "Critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    }

    # Apply the color to the 'Vulnerability' column (assumed to be column 'I' or the 9th column)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
        for cell in row:
            cell_value = cell.value
            if cell_value in color_mapping:
                cell.fill = color_mapping[cell_value]

    wb.save(excel_file)


# Example Usage
# urls_to_check = [
#     "http://example-site.com/_login?user=admin;session=123",
#     "https://w6xvbucu6.top/",
#     "http://example.com:8080/home",
#     "https://R2sBBM0e.go.jp@jbjampcejm.sbs/FisCNm",
#     "http://valid-site.com/user_123",
#     "http://testsafebrowsing.appspot.com/s/malware.html",
#     "http://0xC0A80101",
#     "http://192.168.1.1",
#     "http://192.168.168.88/",
#     "http://themestarz.net/html/appstorm/?storefront=envato-elements",
#     "https://elements.envato.com/vniapp-showcase-mobile-app-html-template-FY84QW7",
#     "https://www.google.com",
#     "https://www.amazon.com",
#     "https://www.growtharc.com/",
#     "www.auxoai.com",
#     "https://www.relanto.ai/",
#     "https://www.accenture.com/in-en",
# ]



# # Analyzing the URLs and getting the DataFrame
# url_scores_df = analyze_urls(urls_to_check)

# # Save the DataFrame to an Excel file
# output_file = "url_analysis_results.xlsx"
# url_scores_df.to_excel(output_file, index=False)

# # Apply color coding to the vulnerability column
# apply_color_to_vulnerability_column(output_file, "Sheet1")

# print(f"Results have been saved to {output_file}")